import folium
from folium.plugins import MarkerCluster
import openpyxl as oxl
import xlrd
import datetime


def read_gas_list(file_name):
    # Расширение файла
    file_ext = file_name.split(".")[-1]

    if file_ext == "xls":
        workb = xlrd.open_workbook(file_name)
        # Загружаем активный лист
        ws = workb.sheet_by_index(0)
        # Получаем максимальное количество заполненных строк
        m_rows = ws.nrows
        xlsx_col = -1  # В openpyxl отсчет столбцов начинается с 1, в xlrd с 0
    elif file_ext == "xlsx":
        workb = oxl.load_workbook(filename=file_name, read_only=True, data_only=True)
        # Загружаем активный лист
        ws = workb.active
        # Получаем максимальное количество заполненных строк
        m_rows = ws.max_row
        xlsx_col = 0  # В openpyxl отсчет столбцов начинается с 1, в xlrd с 0

    brand = []
    longitude = []
    latitude = []
    address = []
    diesel = []

    # Пробегаем файл по строчкам
    cur_row = 5
    while cur_row < m_rows - 10:
        if cur_row % 100 == 0:
            print(cur_row)

        lon = ws.cell(cur_row, 9 + xlsx_col).value
        lat = ws.cell(cur_row, 10 + xlsx_col).value
        ai92 = ws.cell(cur_row, 13 + xlsx_col).value
        ai95 = ws.cell(cur_row, 14 + xlsx_col).value
        if lon == 0 or lat == 0 or lon == [] or lat == [] \
                or ai92 == "-" or ai95 == "-":
            cur_row += 1
            continue
        else:
            longitude.append(lon)
            latitude.append(lat)

        brand.append(ws.cell(cur_row, 8 + xlsx_col).value)
        address.append(ws.cell(cur_row, 20 + xlsx_col).value)
        diesel.append(True if ws.cell(cur_row, 11 + xlsx_col).value == "+" else False)

        cur_row += 1

    return brand, longitude, latitude, address, diesel


def create_map(brand, longitude, latitude, address, diesel):
    # Latitude - долгота
    # Longitude - широта

    # Create base map
    mp = folium.Map(location=[54.9893, 82.9070],
                    zoom_start=7,
                    tiles='OpenStreetMap')
    # Other tiles: Stamen Terrain, Stamen Toner, Mapbox Bright, Mapbox Control Room and other

    # Create Cluster
    marker_cluster = MarkerCluster().add_to(mp)

    for i in range(len(longitude)):
        if longitude[i] >= 56:                          # Сохраняем Сибирь и Восток
            # Plot markers and add to marker_cluster
            icon_name = 'truck' if diesel[i] else 'car'

            folium.Marker(location=[latitude[i], longitude[i]],
                          popup=address[i],
                          tooltip=brand[i],
                          icon=folium.Icon(icon=icon_name, prefix='fa')).add_to(marker_cluster)
            # more icons on https://fontawesome.com/v4.7.0/icons/

    # Enable lat/lon popovers
    mp.add_child(folium.LatLngPopup())

    # Save the map
    mp.save("map1.html")


if __name__ == "__main__":
    start1 = datetime.datetime.now()
    brand, longitude, latitude, address, diesel = read_gas_list("gas list.xls")
    print("Время выполнения 1: " + str(datetime.datetime.now() - start1))
    print(len(latitude))

    start2 = datetime.datetime.now()
    create_map(brand, longitude, latitude, address, diesel)
    print("Время выполнения 2: " + str(datetime.datetime.now() - start2))
